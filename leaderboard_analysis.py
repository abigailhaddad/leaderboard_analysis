import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
import dateparser
import matplotlib.pyplot as plt

def extract_hyperlinks(file_path, sheet_name, column_name):
    """
    Extract hyperlinks from a specified column in an Excel file and returns a dataframe
    containing all data from the Excel file, with hyperlinks in the specified column replaced
    with the hyperlink targets.
    """
    ...

    # Load your workbook
    wb = load_workbook(filename=file_path)

    # Select the worksheet that has your data
    ws = wb[sheet_name]

    # Find the column number of the hyperlink column
    column_num = None
    for column in ws.iter_cols(1, ws.max_column):
        if column[0].value == column_name:
            column_num = column[0].col_idx
            break

    if column_num is None:
        raise ValueError(f"No column named {column_name} in the worksheet")

    # Create two lists to hold your link text and URLs
    link_text = []
    link_url = []

    # Go through each cell in the hyperlink column
    for cell in ws.iter_rows(min_row=2, min_col=column_num, max_col=column_num, max_row=ws.max_row):
        for cell_item in cell:
            if cell_item.hyperlink:
                link_text.append(cell_item.value)
                link_url.append(cell_item.hyperlink.target)

    # Create a DataFrame
    df_links = pd.DataFrame({
        column_name: link_text,
        'URL': link_url
    })

    # Read the existing data into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # Drop Initial column
    df = df.drop(columns=[column_name])
    # Concatenate the two dataframes
    df = pd.concat([df, df_links], axis=1)

    return df

def get_page_text(url):
    """
    Sends a GET request to the GitHub commits page corresponding to the provided url,
    and returns the text content of the page.
    """

    gitURL= url + "/commits/main"
    response = requests.get(gitURL)
    
    # If the GET request is successful, the status code will be 200
    if response.status_code != 200:
        return None

    # Get the content of the response
    page_content = response.content

    # Create a BeautifulSoup object and specify the parser
    soup = BeautifulSoup(page_content, 'html.parser')

    # Extract the text from the page
    text = soup.get_text()

    return text


def extract_dates(text):
    # Regular expression pattern for "committed on" followed by anything
    pattern = r'commited on\s+(.+)'

    # Find all matches in the text
    matches = re.findall(pattern, text)

    # Here matches will contain date strings or relative time strings like '10 days ago'
    return matches


def pull_dates(row):
    text = row['text']
    url = row['URL']
    # If the text is None, return the URL
    if text is None:
        return url
    dates = extract_dates(text)
    # If no dates were found, return the URL instead
    if not dates:
        return url
    return dates



def standardize_dates(date_list):
    standardized_dates = []
    current_year = datetime.now().year
    for date_str in date_list:
        # Parse the date string
        date = dateparser.parse(date_str)

        # If the year is missing, set it to the current year
        if date.year is None:
            date = date.replace(year=current_year)

        # Add the standardized date to the list
        standardized_dates.append(date)

    # Sort the list of dates and return the first one
    standardized_dates.sort()
    return standardized_dates[0]


def restructure_dataframe(df, start_date):
    # Ensure that 'first_commit_date' is datetime
    df['first_commit_date'] = pd.to_datetime(df['first_commit_date'])
    
    # Create a new DataFrame with dates from start_date to current date as index
    date_range = pd.date_range(start=start_date, end=pd.to_datetime('today'), freq='D')
    df_new = pd.DataFrame(index=date_range)

    # For each date, get the best available scores from df where 'first_commit_date' is on or before the date
    score_cols = ['Average ', 'ARC 25s ', 'HellaSwag 10s ', 'MMLU 5s ', 'TruthfulQA MC 0s ']
    for date in df_new.index:
        df_temp = df[df['first_commit_date'] <= date]
        for col in score_cols:
            max_score = df_temp[col].max()
            df_new.at[date, col] = max_score
            if not pd.isna(max_score):
                # Get the name of the model associated with the max score
                model_name = df_temp[df_temp[col] == max_score]['Model'].values[0]
                df_new.at[date, f'{col}_model'] = model_name

    return df_new


def plot_scores(df, df_gpt, label_min_days=15):
    fig, axs = plt.subplots(2, 2, figsize=(15, 10))
    fig.suptitle('HuggingFace Leaderboard Scores vs. GPT-3.5, GPT-4', fontsize=16)

    score_cols = ['ARC 25s ', 'HellaSwag 10s ', 'MMLU 5s ', 'TruthfulQA MC 0s ']
    gpt_colors = {'GPT-3.5': 'g', 'GPT-4': 'b'}

    for ax, col in zip(axs.flat, score_cols):
        if col in df_gpt.columns:
            for gpt in df_gpt.index:
                start_time = pd.to_datetime(df_gpt.at[gpt, 'startingtime'])
                if start_time < df.index[0]:  # if the start_time is before the first date in df, adjust start_time to the first date in df
                    start_time = df.index[0]
                end_time = df.index[-1]  # use the last date in df as the end time
                x = pd.date_range(start=start_time, end=end_time, freq='D')  # create a date range from start_time to end_time
                y = [df_gpt.at[gpt, col]] * len(x)  # create a list of the score repeated for each date in x
                ax.plot(x, y, color=gpt_colors[gpt], linestyle='--')

                # Add label to the line
                midpoint = start_time + (end_time - start_time) / 2
                ax.text(midpoint, df_gpt.at[gpt, col], gpt, ha='center')

        df[col].dropna().plot(ax=ax)
        ax.set_title(col)

        # Add model name labels to horizontal lines longer than label_min_days
        prev_val = None
        prev_date = None
        last_date = df.index[-1]  # get the last date
        for date, val in df[col].items():
            if pd.isna(val):
                continue
            if prev_val is None:
                prev_val = val
                prev_date = date
            elif val != prev_val or date == last_date:  # check if it's the last date
                if (date - prev_date).days >= label_min_days or date == last_date:  # add condition to label the line if it's the last date
                    # Add a label at the midpoint of the line
                    midpoint = prev_date + (date - prev_date) / 2
                    # Get closest date in the index to the midpoint
                    closest_date = df.index[df.index.get_loc(midpoint, method='nearest')]
                    model_name = df.at[closest_date, f'{col}_model']
                    ax.text(midpoint, prev_val, model_name, ha='center')
                prev_val = val
                prev_date = date

    plt.tight_layout()
    plt.savefig("leaderboard_scores.png")
    plt.show()




def create_gpt_df():
    data = {
        'startingtime': ['2022-11-28', '2023-03-14'],
        'ARC 25s ': [85.2, 96.3],  # Replace these with actual GPT-3.5 and GPT-4 scores
        'HellaSwag 10s ': [85.5, 95.3],  # Replace these with actual GPT-3.5 and GPT-4 scores
        'MMLU 5s ': [70.0, 86.4],  # Replace these with actual GPT-3.5 and GPT-4 scores
    }
    df_gpt = pd.DataFrame(data, index=['GPT-3.5', 'GPT-4'])
    return df_gpt




def run_leaderboard_analysis(file_path, sheet_name, column_name, start_date):
    df = extract_hyperlinks(file_path, sheet_name, column_name)
    print(df.loc[~df['URL'].astype(str).str.contains("https:")]['Model'].unique())
    df=df.loc[df['URL'].astype(str).str.contains("https:")]
    df['text']=df['URL'].apply(get_page_text)
    df['dates'] = df.apply(pull_dates, axis=1)
    print(df.loc[df['dates'].astype(str).str.contains("https:")]['Model'].unique())
    df=df.loc[~df['dates'].astype(str).str.contains("https:")]
    df.columns = df.columns.str.replace('[^a-zA-Z0-9\s]', '', regex=True)
    df['first_commit_date']=df['dates'].apply(standardize_dates)
    df_new = restructure_dataframe(df, start_date)
    df_gpt = create_gpt_df()
    # Call the function with your DataFrame
    plot_scores(df_new, df_gpt)

# Now you can run the analysis by calling this function
run_leaderboard_analysis("raw_leaderboard.xlsx", "Sheet1", "Model", "2023-01-01")

